#!/usr/bin/env python3
"""
convert_to_tooth_labels.py

Converts Key Points Annotations JSON bounding boxes to per-image tooth-number label files.

Features:
- Reads metadata from "Characteristics of radiographs included.xlsx" to map image id -> list of FDI teeth.
- Iterates Training/ Testing/ Validation `Key Points Annotations/` JSON files and writes per-image .txt files.
- Supports `raw` mode (FDI xmin ymin xmax ymax) and `yolo` mode (FDI_or_class x_center_norm y_center_norm w_norm h_norm).
- Optionally remaps FDI numbers to zero-based class indices and writes `classes.txt`.

Usage examples:
  python convert_to_tooth_labels.py --dataset-root "/Users/tayebekavousi/Desktop/Dataset" --mode both

"""
import argparse
import json
import math
import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook
    from PIL import Image
except Exception as e:
    raise SystemExit("Missing dependency: run `pip install -r requirements.txt`\n" + str(e))


def find_excel(dataset_root: Path) -> Optional[Path]:
    # Check for txt file first
    txt_file = dataset_root / "characteristics_of_distributions.txt"
    if txt_file.exists():
        return txt_file
    candidates = [dataset_root / "Characteristics of radiographs included.xlsx"]
    for c in candidates:
        if c.exists():
            return c
    # fallback: search
    for p in dataset_root.rglob('*.xlsx'):
        if 'Characteristics' in p.name:
            return p
    return None


def parse_excel(excel_path: Path) -> Dict[str, List[str]]:
    # Check if it's a txt file
    if excel_path.suffix == '.txt':
        mapping: Dict[str, List[str]] = {}
        with open(excel_path, 'r') as f:
            lines = f.readlines()
            # Skip header
            for line in lines[1:]:
                line = line.strip()
                if not line:
                    continue
                parts = line.split('\t')
                if len(parts) < 4:
                    continue
                img_id = parts[0].strip()
                fdi_str = parts[3].strip()
                if not fdi_str:
                    mapping[img_id] = []
                    continue
                fdi_list = re.split(r'[;,\s]+', fdi_str)
                fdi_list = [f for f in fdi_list if f]
                mapping[img_id] = fdi_list
        # Filter out specific IDs to exclude
        exclude_ids = {'863', '777', '762', '75'}
        mapping = {k: v for k, v in mapping.items() if k not in exclude_ids}
        return mapping
    
    # Original Excel parsing
    wb = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
    ws = wb.active
    # read header
    header = []
    first_row = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        first_row = row
        break
    if first_row is None:
        raise RuntimeError('Excel file is empty')
    header = [str(c).strip() if c is not None else '' for c in first_row]
    cols = {c.lower(): i for i, c in enumerate(header)}

    id_col_idx = None
    for candidate in ('id', 'image id', 'image_id', 'img_id'):
        if candidate in cols:
            id_col_idx = cols[candidate]
            break
    if id_col_idx is None:
        id_col_idx = 0

    fdi_col_idx = None
    for k, idx in cols.items():
        if 'fdi' in k or 'tooth' in k or 'teeth' in k or 'notation' in k:
            fdi_col_idx = idx
            break
    if fdi_col_idx is None:
        raise RuntimeError('Could not find an FDI/teeth column in the Excel file')

    mapping: Dict[str, List[str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_id = row[id_col_idx] if id_col_idx < len(row) else None
        if raw_id is None:
            continue
        if isinstance(raw_id, (int, float)) and not math.isnan(raw_id):
            img_id = str(int(raw_id))
        else:
            img_id = str(raw_id)
        raw_fdi = row[fdi_col_idx] if fdi_col_idx < len(row) else None
        if raw_fdi is None:
            mapping[img_id] = []
            continue
        s = str(raw_fdi)
        parts = re.split(r'[;,\s]+', s.strip())
        parts = [p for p in parts if p]
        mapping[img_id] = parts
    wb.close()
    return mapping


def looks_like_box_list(obj: Any) -> bool:
    # check for list of lists of 4 numeric values
    if not isinstance(obj, list):
        return False
    if not obj:
        return False
    for el in obj:
        if not isinstance(el, (list, tuple)):
            return False
        if len(el) != 4:
            return False
        if not all(isinstance(x, (int, float)) for x in el):
            return False
    return True


def recursive_find_boxes(obj: Any) -> Optional[List[List[float]]]:
    if looks_like_box_list(obj):
        # convert to float lists
        return [[float(x) for x in b] for b in obj]
    if isinstance(obj, dict):
        # common keys
        for k in ('data', 'annotations', 'boxes', 'bboxes', 'bounding_boxes', 'bbox'):
            if k in obj:
                res = recursive_find_boxes(obj[k])
                if res:
                    return res
        # otherwise traverse
        for v in obj.values():
            res = recursive_find_boxes(v)
            if res:
                return res
    if isinstance(obj, list):
        for v in obj:
            res = recursive_find_boxes(v)
            if res:
                return res
    return None


def find_image_file(images_dir: Path, img_id: str) -> Optional[Path]:
    # try common extensions
    for ext in ('.jpg', '.jpeg', '.png', '.tif', '.tiff', '.bmp'):
        p = images_dir / (img_id + ext)
        if p.exists():
            return p
    # try any file starting with img_id
    if images_dir.exists():
        for p in images_dir.iterdir():
            if p.stem == img_id:
                return p
    return None


def get_image_size_from_json_or_file(j: Dict[str, Any], images_dir: Path, img_id: str) -> Optional[Tuple[int, int]]:
    # try common keys in json
    if isinstance(j, dict):
        if 'width' in j and 'height' in j and isinstance(j['width'], (int, float)) and isinstance(j['height'], (int, float)):
            return int(j['width']), int(j['height'])
    # some jsons nest image info
    for k in ('image', 'meta'):
        if k in j and isinstance(j[k], dict):
            sub = j[k]
            if 'width' in sub and 'height' in sub:
                return int(sub['width']), int(sub['height'])
    # fallback to open image file
    img_path = find_image_file(images_dir, img_id)
    if img_path is None:
        return None
    try:
        with Image.open(img_path) as im:
            return im.width, im.height
    except Exception:
        return None


def to_yolo(box: List[float], W: int, H: int) -> Tuple[float, float, float, float]:
    xmin, ymin, xmax, ymax = box
    x_c = (xmin + xmax) / 2.0
    y_c = (ymin + ymax) / 2.0
    w = xmax - xmin
    h = ymax - ymin
    return x_c / W, y_c / H, w / W, h / H


def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def process_split(split_dir: Path, metadata: Dict[str, List[str]], label_dir_name: str, mode: str,
                  remap: bool, decimals: int, class_map: Dict[str, int]) -> None:
    kpa_dir = split_dir / 'Key Points Annotations'
    images_dir = split_dir / 'Images'
    out_dir = split_dir / label_dir_name
    ensure_dir(out_dir)
    if not kpa_dir.exists():
        print(f'Warning: {kpa_dir} does not exist, skipping')
        return
    json_files = sorted(kpa_dir.glob('*.json'))
    for jf in json_files:
        img_id = jf.stem
        try:
            j = json.loads(jf.read_text())
        except Exception as e:
            print(f'Failed to parse {jf}: {e}')
            continue
        boxes = None
        # first try j['data'] if present
        if 'data' in j:
            boxes = recursive_find_boxes(j['data'])
        if boxes is None:
            boxes = recursive_find_boxes(j)
        if boxes is None:
            print(f'No bounding-box list found in {jf}, skipping')
            continue

        fdi_list = metadata.get(img_id, [])
        if not fdi_list:
            print(f'No FDI metadata for image {img_id} (file {jf.name}), skipping')
            continue

        if len(fdi_list) != len(boxes):
            print(f'Skipping {img_id}: count mismatch ({len(fdi_list)} FDIs vs {len(boxes)} boxes)')
            continue

        n = len(fdi_list)

        # determine image size if YOLO mode requested
        size = None
        if mode in ('yolo', 'both'):
            size = get_image_size_from_json_or_file(j, images_dir, img_id)
            if size is None:
                print(f'Could not determine image size for {img_id}; skipping')
                continue

        out_lines: List[str] = []
        for i in range(n):
            fdi = fdi_list[i]
            box = boxes[i]
            xmin, ymin, xmax, ymax = [float(x) for x in box]
            # Always write raw coordinates first
            if mode in ('raw', 'both'):
                out_lines.append(f"{fdi} {int(xmin)} {int(ymin)} {int(xmax)} {int(ymax)}")
        
        # Then write YOLO normalized coordinates
        if mode in ('yolo', 'both') and size is not None:
            W, H = size
            if W == 0 or H == 0:
                print(f'Invalid image size for {img_id}: W={W}, H={H}, skipping')
                continue
            for i in range(n):
                fdi = fdi_list[i]
                box = boxes[i]
                xmin, ymin, xmax, ymax = [float(x) for x in box]
                xcn, ycn, wn, hn = to_yolo([xmin, ymin, xmax, ymax], W, H)
                if remap:
                    cls = class_map.get(str(fdi))
                    if cls is None:
                        # this should not happen if class_map built from metadata
                        cls = 0
                    first = str(cls)
                else:
                    first = str(fdi)
                out_lines.append(f"{first} {xcn:.{decimals}f} {ycn:.{decimals}f} {wn:.{decimals}f} {hn:.{decimals}f}")

        # write output file
        out_path = out_dir / (img_id + '.txt')
        out_path.write_text('\n'.join(out_lines) + ('\n' if out_lines else ''))


def build_class_map(metadata: Dict[str, List[str]]) -> Dict[str, int]:
    s = set()
    for v in metadata.values():
        for f in v:
            s.add(str(f))
    sorted_f = sorted(s, key=lambda x: (0, int(x)) if x.isdigit() else (1, x))
    return {f: i for i, f in enumerate(sorted_f)}


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--dataset-root', default='.', help='Path to dataset root (contains Training Testing Validation)')
    p.add_argument('--mode', choices=('raw', 'yolo', 'both'), default='both', help='Which outputs to write')
    p.add_argument('--label-dir-name', default='Tooth_Labels', help='Name of label directory to create in each split')
    p.add_argument('--remap-classes', action='store_true', help='Remap FDI numbers to 0-based class ids for YOLO output')
    p.add_argument('--decimals', type=int, default=6, help='Decimal places for YOLO floats')
    args = p.parse_args()

    dataset_root = Path(args.dataset_root)
    excel = find_excel(dataset_root)
    if excel is None:
        raise SystemExit('Could not find "Characteristics of radiographs included.xlsx" under dataset root')
    print('Using metadata:', excel)
    metadata = parse_excel(excel)

    class_map = {}
    if args.remap_classes:
        class_map = build_class_map(metadata)
        # write classes.txt at dataset root
        classes_path = dataset_root / 'classes.txt'
        # ensure sorted by index
        inv = {v: k for k, v in class_map.items()}
        lines = [inv[i] for i in range(len(inv))]
        classes_path.write_text('\n'.join(lines) + '\n')
        print(f'Wrote class mapping to {classes_path}')

    for split in ('Training', 'Testing', 'Validation'):
        split_dir = dataset_root / split
        if not split_dir.exists():
            print(f'Warning: split dir {split_dir} does not exist, skipping')
            continue
        print(f'Processing split {split}')
        process_split(split_dir, metadata, args.label_dir_name, args.mode, args.remap_classes, args.decimals, class_map)

    print('Done.')


if __name__ == '__main__':
    main()
